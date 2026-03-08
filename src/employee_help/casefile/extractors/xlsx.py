"""ExcelExtractor — text extraction from .xlsx files as markdown tables."""

from __future__ import annotations

import io

from employee_help.casefile.extractors.base import ExtractionResult, FileExtractor

_SUPPORTED_EXTENSIONS = {"xlsx"}

_SUPPORTED_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def _has_openpyxl() -> bool:
    """Check if openpyxl is available."""
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def _cell_to_str(value: object) -> str:
    """Convert a cell value to a string, escaping pipe characters for markdown."""
    if value is None:
        return ""
    text = str(value).replace("|", "\\|").replace("\n", " ")
    return text


def _sheet_to_markdown(sheet) -> str:  # type: ignore[no-untyped-def]
    """Convert an openpyxl worksheet to a markdown table.

    Skips entirely empty rows and columns. Returns empty string if
    the sheet has no data.
    """
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_cell_to_str(v) for v in row])

    if not rows:
        return ""

    # Trim trailing empty rows
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()

    if not rows:
        return ""

    # Find max column count (some rows may be shorter)
    max_cols = max(len(r) for r in rows)
    if max_cols == 0:
        return ""

    # Pad shorter rows
    for r in rows:
        while len(r) < max_cols:
            r.append("")

    # Trim trailing empty columns
    while max_cols > 0 and all(r[max_cols - 1] == "" for r in rows):
        for r in rows:
            r.pop()
        max_cols -= 1

    if max_cols == 0:
        return ""

    # Build markdown table
    lines: list[str] = []

    # Header row
    header = rows[0]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")

    # Data rows
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


class ExcelExtractor(FileExtractor):
    """Extract text from Excel (.xlsx) files as markdown tables.

    Each sheet is rendered as a separate markdown table with the sheet
    name as a heading. Requires the ``openpyxl`` package (part of the
    ``[casefile]`` dependency group).
    """

    def can_extract(self, mime_type: str, extension: str) -> bool:
        return extension in _SUPPORTED_EXTENSIONS or mime_type in _SUPPORTED_MIMES

    @property
    def supported_extensions(self) -> set[str]:
        return set(_SUPPORTED_EXTENSIONS)

    def extract(self, file_bytes: bytes, filename: str) -> ExtractionResult:
        warnings: list[str] = []

        if not _has_openpyxl():
            return ExtractionResult(
                text="",
                metadata={"extractor": "xlsx"},
                warnings=["openpyxl is not installed; cannot extract Excel files"],
            )

        import openpyxl

        if not file_bytes:
            return ExtractionResult(
                text="",
                metadata={"extractor": "xlsx", "sheet_count": 0, "sheet_names": []},
                warnings=["No text extracted from document"],
            )

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            return ExtractionResult(
                text="",
                metadata={"extractor": "xlsx"},
                warnings=[f"Failed to open Excel file: {exc}"],
            )

        sections: list[str] = []
        sheet_names: list[str] = []
        total_rows = 0

        try:
            for sheet_name in wb.sheetnames:
                sheet_names.append(sheet_name)
                ws = wb[sheet_name]

                table_md = _sheet_to_markdown(ws)
                if not table_md:
                    warnings.append(f"Sheet '{sheet_name}' is empty")
                    continue

                row_count = table_md.count("\n")  # approximate
                total_rows += row_count

                if len(wb.sheetnames) > 1:
                    sections.append(f"## {sheet_name}\n\n{table_md}")
                else:
                    sections.append(table_md)
        finally:
            wb.close()

        text = "\n\n".join(sections)

        if not text.strip():
            warnings.append("No text extracted from document")

        return ExtractionResult(
            text=text,
            metadata={
                "extractor": "xlsx",
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "total_data_rows": total_rows,
            },
            warnings=warnings,
        )
